import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def capitalize_excel(input_file, output_file):
    # Load the workbook and modify in place
    wb = load_workbook(input_file, data_only=False)
    
    # Process each sheet without losing formatting or images
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.upper()
    
    # Save the modified file
    wb.save(output_file)
    print(f"File saved as {output_file} ?")
    )

def select_file():
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(title="Select an Excel File", filetypes=[("Excel files", "*.xlsx;*.xlsm")])
    return file_path

# Open file explorer to choose input file
input_excel = select_file()
if input_excel:
    # Extract file name and create output path in Windows Downloads folder
    base_name = os.path.basename(input_excel)
    name, ext = os.path.splitext(base_name)
    output_excel = os.path.join(os.path.expanduser("~"), "Downloads", f"{name}-cap{ext}")
    
    capitalize_excel(input_excel, output_excel)
else:
    print("No file selected.")
